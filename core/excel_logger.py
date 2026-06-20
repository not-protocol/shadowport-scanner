"""
core/excel_logger.py — ShadowPort Scanner v2.3.0
Dual-write Excel logger — writes immediately after every scan/plugin.
Thread-safe. Auto-creates Log/ directory and file.
"""

import os
import threading
import time
from datetime import datetime
from typing import Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config.settings import EXCEL_PATH, LOG_DIR

_lock = threading.Lock()

_SCAN_HEADERS   = ["Scan #","Date","Time","Target","Scan Type","Open Ports","Total Ports","Risk Score","Duration (s)","Status"]
_PLUGIN_HEADERS = ["#","Date","Time","Target","Plugin","Output Preview","Duration (s)","Status"]

_H_FILL  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_H_FONT  = Font(bold=True, color="FFFFFF", size=11)
_H_ALIGN = Alignment(horizontal="center", vertical="center")


def _styled_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for i, cell in enumerate(ws[1], 1):
        cell.font, cell.fill, cell.alignment = _H_FONT, _H_FILL, _H_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(headers[i-1]) + 4)
    ws.row_dimensions[1].height = 20


def _load_or_create(path: str) -> Workbook:
    if os.path.exists(path):
        try:
            return load_workbook(path)
        except Exception:
            pass
    wb = Workbook()
    wb.active.title = "Scans"
    wb.create_sheet("Plugins")
    _styled_header(wb["Scans"],   _SCAN_HEADERS)
    _styled_header(wb["Plugins"], _PLUGIN_HEADERS)
    return wb


def _next_num(ws) -> int:
    return max(ws.max_row - 1, 0) + 1


def _save_with_retry(wb: Workbook, path: str) -> Tuple[bool, str]:
    for attempt in range(1, 4):
        try:
            wb.save(path)
            wb.close()
            return True, ""
        except PermissionError as e:
            if attempt < 3:
                time.sleep(1.0)
            else:
                return False, str(e)
        except Exception as e:
            return False, str(e)
    return False, "Unknown error"


def log_scan_to_excel(scan_data: dict, risk_score: int = 0) -> Tuple[bool, str]:
    os.makedirs(LOG_DIR, exist_ok=True)
    now    = datetime.now()
    ports  = scan_data.get("ports", [])
    open_c = sum(1 for p in ports if p.get("state") == "open")

    with _lock:
        wb = _load_or_create(EXCEL_PATH)
        ws = wb["Scans"] if "Scans" in wb.sheetnames else wb.active
        if ws.max_row == 0 or ws.cell(1,1).value != "Scan #":
            ws.delete_rows(1, ws.max_row)
            _styled_header(ws, _SCAN_HEADERS)

        ws.append([
            _next_num(ws),
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            scan_data.get("host",""),
            scan_data.get("mode_name",""),
            open_c,
            len(ports),
            int(risk_score),
            round(float(scan_data.get("duration_seconds", 0.0)), 1),
            scan_data.get("state","unknown"),
        ])
        return _save_with_retry(wb, EXCEL_PATH)


def log_plugin_to_excel(target: str, plugin_name: str,
                        output: str, duration: float,
                        success: bool) -> Tuple[bool, str]:
    os.makedirs(LOG_DIR, exist_ok=True)
    now = datetime.now()
    preview = (output[:80] + "…") if len(output) > 80 else output

    with _lock:
        wb = _load_or_create(EXCEL_PATH)
        if "Plugins" not in wb.sheetnames:
            ws = wb.create_sheet("Plugins")
            _styled_header(ws, _PLUGIN_HEADERS)
        else:
            ws = wb["Plugins"]

        ws.append([
            _next_num(ws),
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            target,
            plugin_name,
            preview,
            round(duration, 2),
            "Success" if success else "Failed",
        ])
        return _save_with_retry(wb, EXCEL_PATH)
