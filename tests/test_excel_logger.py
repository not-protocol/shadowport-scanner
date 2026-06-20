"""tests/test_excel_logger.py — ShadowPort Scanner v2.3.0"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pytest
from openpyxl import load_workbook
import core.excel_logger as excel_mod
from core.excel_logger import log_scan_to_excel, log_plugin_to_excel, _SCAN_HEADERS, _PLUGIN_HEADERS


SCAN = {
    "host": "192.168.1.1", "mode_name": "Quick Scan", "state": "up",
    "duration_seconds": 12.5,
    "ports": [
        {"port":"22","proto":"tcp","state":"open","service":"ssh"},
        {"port":"80","proto":"tcp","state":"open","service":"http"},
        {"port":"443","proto":"tcp","state":"closed","service":"https"},
    ],
    "risk": {"score": 14, "label": "LOW EXPOSURE", "breakdown": []},
}


@pytest.fixture(autouse=True)
def temp_excel(tmp_path, monkeypatch):
    path = str(tmp_path / "scannerhistory.xlsx")
    monkeypatch.setattr(excel_mod, "EXCEL_PATH", path)
    monkeypatch.setattr(excel_mod, "LOG_DIR",    str(tmp_path))
    import config.settings as s
    monkeypatch.setattr(s, "EXCEL_PATH", path)
    monkeypatch.setattr(s, "LOG_DIR",    str(tmp_path))


# ── new file ──────────────────────────────────────────────────────────────────

def test_creates_file():
    ok, msg = log_scan_to_excel(SCAN, risk_score=14)
    assert ok, msg
    assert os.path.exists(excel_mod.EXCEL_PATH)

def test_has_scans_and_plugins_sheets():
    log_scan_to_excel(SCAN, risk_score=14)
    wb = load_workbook(excel_mod.EXCEL_PATH)
    assert "Scans" in wb.sheetnames
    assert "Plugins" in wb.sheetnames
    wb.close()

def test_scan_header_correct():
    log_scan_to_excel(SCAN, risk_score=14)
    wb = load_workbook(excel_mod.EXCEL_PATH)
    ws = wb["Scans"]
    headers = [ws.cell(1,i).value for i in range(1, len(_SCAN_HEADERS)+1)]
    assert headers == _SCAN_HEADERS
    wb.close()

def test_first_scan_number_is_1():
    log_scan_to_excel(SCAN, risk_score=14)
    wb = load_workbook(excel_mod.EXCEL_PATH)
    assert wb["Scans"].cell(2,1).value == 1
    wb.close()


# ── append ────────────────────────────────────────────────────────────────────

def test_append_increments():
    log_scan_to_excel(SCAN, risk_score=14)
    log_scan_to_excel(SCAN, risk_score=14)
    wb = load_workbook(excel_mod.EXCEL_PATH)
    ws = wb["Scans"]
    assert ws.cell(2,1).value == 1
    assert ws.cell(3,1).value == 2
    assert ws.max_row == 3
    wb.close()

def test_scan_row_values():
    log_scan_to_excel(SCAN, risk_score=14)
    wb = load_workbook(excel_mod.EXCEL_PATH)
    ws  = wb["Scans"]
    row = [ws.cell(2,i).value for i in range(1, len(_SCAN_HEADERS)+1)]
    assert row[3] == "192.168.1.1"
    assert row[5] == 2     # open ports
    assert row[6] == 3     # total ports
    assert row[7] == 14    # risk
    wb.close()


# ── plugin logging ────────────────────────────────────────────────────────────

def test_plugin_log_creates_row():
    ok, msg = log_plugin_to_excel("192.168.1.1", "dns_lookup", "Forward: 1.2.3.4", 0.5, True)
    assert ok, msg
    wb = load_workbook(excel_mod.EXCEL_PATH)
    ws = wb["Plugins"]
    headers = [ws.cell(1,i).value for i in range(1, len(_PLUGIN_HEADERS)+1)]
    assert headers == _PLUGIN_HEADERS
    assert ws.cell(2,1).value == 1
    assert ws.cell(2,4).value == "192.168.1.1"
    assert ws.cell(2,5).value == "dns_lookup"
    assert ws.cell(2,8).value == "Success"
    wb.close()

def test_plugin_failure_status():
    log_plugin_to_excel("10.0.0.1", "broken", "err", 0.1, False)
    wb = load_workbook(excel_mod.EXCEL_PATH)
    ws = wb["Plugins"]
    assert ws.cell(2,8).value == "Failed"
    wb.close()

def test_plugin_output_preview_truncated():
    long_output = "x" * 200
    log_plugin_to_excel("10.0.0.1", "test", long_output, 0.1, True)
    wb = load_workbook(excel_mod.EXCEL_PATH)
    ws = wb["Plugins"]
    val = ws.cell(2,6).value
    assert len(val) <= 81  # 80 chars + ellipsis
    wb.close()


# ── missing dir ───────────────────────────────────────────────────────────────

def test_missing_dir_autocreated(tmp_path, monkeypatch):
    deep = str(tmp_path / "a" / "b" / "Log")
    path = os.path.join(deep, "scannerhistory.xlsx")
    monkeypatch.setattr(excel_mod, "EXCEL_PATH", path)
    monkeypatch.setattr(excel_mod, "LOG_DIR",    deep)
    ok, _ = log_scan_to_excel(SCAN, risk_score=0)
    assert ok
    assert os.path.exists(path)
